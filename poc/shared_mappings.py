"""
Append-only shared learned-mappings log, kept in a real .xlsx file so it's
directly readable/reviewable in Excel by anyone -- unlike db.py's SQLite
mappings.db, which is opaque outside this tool. Every confirmation appends a
new timestamped row rather than editing one in place: it can never silently
overwrite someone else's row the way an update-in-place design could.

Writes are batched, not immediate (Michael's call, 2026-08-16, after the
per-confirmation-write design turned out to have two real problems at scale:
every confirmation did a full read-modify-write of the whole file, getting
slower as the log grows, and app.py's own stop.bat hard-kills the process
-- Stop-Process -Force -- so there was never a reliable "flush on exit"
moment to depend on either). Confirmations are queued locally instead (see
db.py's pending_sync table) and flushed here in batches by app.py's
background thread, every few minutes or every few confirmations, whichever
comes first -- bounding both the per-write cost and how much could ever be
at risk from an ungraceful exit to "a few minutes," not "the whole session."

flush_pending() is the write path now (append()/append_value_map() are
gone -- nothing outside this file called them once app.py switched to
queuing). Its protocol, in order:
  1. A best-effort lock file (_acquire_lock) -- narrows the collision
     window between two machines flushing at nearly the same moment, but
     is NOT a real lock: OneDrive syncs each machine's own local replica
     of this folder asynchronously, so a lock file written here has no
     immediate effect on another machine's view of it. Treat it as
     "usually helps," not "guarantees exclusivity."
  2. Read the file fresh from disk (not any cached in-memory state) to see
     which of the pending rows' SyncID values are already present --
     already-present ones are a retried flush confirming a row that
     really did make it in on a prior attempt whose *verification* step
     is what failed, not the write itself. They're reported back as
     confirmed without being written again.
  3. Append everything else in ONE write.
  4. Read the file back and verify those rows are actually there now.
  5. Release the lock.
flush_pending() returns exactly the pending rows now confirmed present in
the file -- db.py's clear_pending_sync should only be called with those.
Anything not returned stays queued; the next flush cycle retries it. A row
is never dropped from the local queue on the strength of "the write was
attempted" -- only on "the write was verified."

db.py is untouched by any of this and stays exactly as it is -- a fast,
always-available per-machine cache. This module is a second, independent
read+write path that app.py consults alongside it (see poc/README.md,
"Shared mapping log").

Optional dependency: openpyxl. Only imported if CW_ETL_SHARED_XLSX is set --
an installation that doesn't use this feature never needs it.
"""

import datetime
import getpass
import os
import socket
import time

from db import normalize

SHARED_XLSX_PATH = os.environ.get("CW_ETL_SHARED_XLSX")

# (column header written to the sheet, key used in an in-memory row dict).
# Single source of truth for both directions -- reading builds a row dict
# keyed this way, writing emits values in this same header order.
FIELDS = [
    ("TargetDatabase", "targetDatabase"),
    ("SourceSystem", "sourceSystem"),
    ("SourceField", "sourceField"),
    ("TargetTable", "targetTable"),
    ("TargetField", "targetField"),
    ("ConfirmedAt", "confirmedAt"),
    ("ConfirmedBy", "confirmedBy"),
    # Added after the log already had real rows -- always append new fields
    # here at the end, never insert in the middle. Existing rows' cells stay
    # in their original physical columns; inserting a header in the middle
    # would misalign them against a header that no longer matches where
    # their values actually sit.
    ("Description", "desc"),
    # Same append-at-the-end rule. Holds the customer's confirmed source-
    # value -> target-code mapping from the Advanced-mode value-matching
    # step (e.g. "M=1,F=2,U=3"), written by a value_map-kind flush row --
    # a distinct confirmation from the field-mapping one above, so most
    # rows leave this blank until/unless that step is confirmed for that
    # field.
    ("ValueMap", "valueMap"),
    # Same append-at-the-end rule again. This row's identity in BOTH this
    # file and db.py's pending_sync table -- what flush_pending's
    # idempotent-retry check and verify-after-write step key on. Rows
    # written before this column existed have no SyncID and simply can
    # never collide with one (uuid4 is globally unique), so no migration
    # is needed for them.
    ("SyncID", "syncId"),
]
REQUIRED_KEYS = ("targetDatabase", "sourceSystem", "sourceField", "targetTable", "targetField")
SHEET_NAME = "ConfirmedMappings"

# A lock file older than this is treated as orphaned from a crashed writer,
# not evidence someone's still actively flushing -- cleared and retried
# rather than backed off from forever.
_LOCK_STALE_SECONDS = 120
_LOCK_WAIT_ATTEMPTS = 3
_LOCK_WAIT_SECONDS = 2

_openpyxl = None
_unavailable_reason = None


def _load_openpyxl():
    """Imported lazily and only once -- most installations never touch this
    feature, so most installations never need openpyxl on disk at all."""
    global _openpyxl, _unavailable_reason
    if _openpyxl is not None or _unavailable_reason is not None:
        return _openpyxl
    try:
        import openpyxl
        _openpyxl = openpyxl
    except ImportError:
        _unavailable_reason = "openpyxl isn't installed (run: pip install openpyxl, or see poc/requirements-optional.txt)"
    return _openpyxl


class SharedLog:
    """In-process cache of the shared file's rows, built once by load() and
    kept current in memory as this process appends its own confirmations --
    avoids re-parsing the whole workbook on every suggestion lookup."""

    def __init__(self):
        self.rows = []             # raw row dicts, oldest first
        self.exact = {}            # (target_db, source_system_norm, field_name_norm) -> {target, count, lastConfirmedAt}
        self.field_index = {}      # (target_db, field_name_norm) -> {(table, field): count}
        self.decode_patterns = {}  # (target_db, target_table, target_field) -> {desc: count}
        self.value_maps = {}       # (target_db, source_system_norm, field_name_norm) -> {valueMap, confirmedAt}
        self.known_sync_ids = set()  # every SyncID already in the file -- flush_pending's idempotent-retry check
        self.path = None
        self.status = "disabled"  # disabled | missing-dependency | not-found | ok | error
        self.detail = ""

    def load(self, path):
        self.path = path
        self.rows = []
        self.exact = {}
        self.field_index = {}
        self.decode_patterns = {}
        self.value_maps = {}
        self.known_sync_ids = set()
        if not path:
            self.status = "disabled"
            self.detail = "CW_ETL_SHARED_XLSX not set"
            return
        openpyxl = _load_openpyxl()
        if openpyxl is None:
            self.status = "missing-dependency"
            self.detail = _unavailable_reason
            return
        if not os.path.exists(path):
            self.status = "not-found"
            self.detail = f"{path} does not exist yet"
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            header = None
            for raw_row in ws.iter_rows(values_only=True):
                if header is None:
                    header = [str(c).strip() if c else "" for c in raw_row]
                    continue
                if not any(raw_row):
                    continue
                self._ingest(dict(zip(header, raw_row)))
            wb.close()
            self.status = "ok"
            self.detail = f"{len(self.rows)} row(s) from {path}"
        except Exception as e:  # noqa: BLE001 -- a shared file we don't control must never be able to crash the server
            self.status = "error"
            self.detail = str(e)

    def _ingest(self, record):
        row = {key: (record.get(header) or "") for header, key in FIELDS}
        if not all(row[k] for k in REQUIRED_KEYS):
            return
        self.rows.append(row)
        if row.get("syncId"):
            self.known_sync_ids.add(row["syncId"])
        # Known minor imprecision: a value-map-kind flush row still has a
        # real targetTable/targetField, so a reload counts it as a second
        # confirmation of that same mapping on top of the field-confirm row
        # that already exists -- confirm_count ends up one higher than the
        # number of times the *field mapping itself* was actually
        # reconfirmed. Doesn't affect which table/field/value_map get_exact
        # returns, only that display counter, so left as-is rather than
        # adding a row-type column to distinguish the two flush kinds.
        self._reindex_row(row)
        self._reindex_value_map(row)

    def _reindex_value_map(self, row):
        """Tracks the latest confirmed value_map per (target_db, source
        system, field) -- last-write-wins, same as how a changed field-
        mapping decision overwrites the running count in _reindex_row,
        except there's no count here: a value map is either the current
        confirmed one or it isn't."""
        if not row.get("valueMap"):
            return
        key = (row["targetDatabase"], normalize(row["sourceSystem"]), normalize(row["sourceField"]))
        self.value_maps[key] = {"valueMap": row["valueMap"], "confirmedAt": row.get("confirmedAt") or ""}

    def _reindex_row(self, row):
        tdb = row["targetDatabase"]
        ss_norm = normalize(row["sourceSystem"])
        fn_norm = normalize(row["sourceField"])
        target = (row["targetTable"], row["targetField"])

        # Exact match: track the running count for whichever target is most
        # recent, resetting if a later confirmation named a different one --
        # mirrors db.py's confirm_count semantics (changing your mind about a
        # mapping starts its count over rather than piling onto the old one).
        key = (tdb, ss_norm, fn_norm)
        existing = self.exact.get(key)
        if existing and existing["target"] == target:
            existing["count"] += 1
            existing["lastConfirmedAt"] = row["confirmedAt"] or existing["lastConfirmedAt"]
        else:
            self.exact[key] = {"target": target, "count": 1, "lastConfirmedAt": row["confirmedAt"]}

        # Field index: every confirmation counts, including repeats of the
        # same mapping -- matches db.py's unconditional field_index increment.
        idx_key = (tdb, fn_norm)
        bucket = self.field_index.setdefault(idx_key, {})
        bucket[target] = bucket.get(target, 0) + 1

        # Decode patterns: what consultants typed as the source-side format
        # note for this exact target field, across any source system --
        # never anything from a source system's real data, just field-format
        # text a human typed in. Feeds transform_draft.py's "established
        # pattern" signal the same way db.get_decode_patterns does locally.
        if row.get("desc"):
            pattern_key = (tdb, row["targetTable"], row["targetField"])
            pattern_bucket = self.decode_patterns.setdefault(pattern_key, {})
            pattern_bucket[row["desc"]] = pattern_bucket.get(row["desc"], 0) + 1

    def get_decode_patterns(self, target_db, target_table, target_field):
        bucket = self.decode_patterns.get((target_db, target_table, target_field))
        if not bucket:
            return []
        return sorted(
            ({"desc": d, "count": c} for d, c in bucket.items()),
            key=lambda r: r["count"],
            reverse=True,
        )

    def get_exact(self, target_db, source_system, field_name):
        entry = self.exact.get((target_db, normalize(source_system), normalize(field_name)))
        if not entry:
            return None
        table, field = entry["target"]
        return {
            "target_table": table,
            "target_field": field,
            "confirm_count": entry["count"],
            "last_confirmed_at": entry["lastConfirmedAt"],
            "value_map": self.get_value_map(target_db, source_system, field_name) or "",
        }

    def get_value_map(self, target_db, source_system, field_name):
        entry = self.value_maps.get((target_db, normalize(source_system), normalize(field_name)))
        return entry["valueMap"] if entry else None

    def get_field_index(self, target_db, field_name):
        bucket = self.field_index.get((target_db, normalize(field_name)))
        if not bucket:
            return []
        return sorted(
            ({"target_table": t, "target_field": f, "count": c} for (t, f), c in bucket.items()),
            key=lambda r: r["count"],
            reverse=True,
        )

    def flush_pending(self, pending_rows):
        """Writes every row in pending_rows (dicts shaped like db.py's
        pending_sync rows -- sync_id, kind, target_db, source_system,
        field_name, target_table, target_field, desc, value_map,
        confirmed_at, confirmed_by) to the shared file, batched into ONE
        read-modify-write, and verifies afterward that they actually landed.
        Returns the subset of pending_rows confirmed present in the file
        afterward -- the caller (app.py) should pass exactly that subset to
        db.clear_pending_sync; anything not returned stays queued for the
        next flush attempt. See this module's own docstring for the full
        protocol and why each step exists.

        Never raises for an ordinary "couldn't write this time" outcome
        (missing dependency, file unreachable, lock contention, an
        unexpected exception mid-write) -- returns [] instead, so a
        transient failure just means "try again next cycle," not a crash
        that takes the background flush thread down with it.
        """
        if not pending_rows:
            return []
        openpyxl = _load_openpyxl()
        if openpyxl is None or not self.path:
            return []
        if not _acquire_lock(self.path):
            return []  # another writer looks active -- back off, next scheduled flush will retry
        try:
            if os.path.exists(self.path):
                wb = openpyxl.load_workbook(self.path)
            else:
                wb = openpyxl.Workbook()
            ws = wb.active
            existing_sync_ids = self._extract_sync_ids(ws)

            already_present = [r for r in pending_rows if r["sync_id"] in existing_sync_ids]
            to_write = [r for r in pending_rows if r["sync_id"] not in existing_sync_ids]
            if not to_write:
                return already_present

            self._prepare_sheet(ws)
            for pending_row in to_write:
                row = self._pending_to_row(pending_row)
                ws.append([row[key] for _, key in FIELDS])
            wb.save(self.path)

            # Verify against what's REALLY on disk now, via a fresh read --
            # not the in-memory `wb` we just saved, in case the save didn't
            # actually take (e.g. a sync client replaced the file with a
            # conflicted copy in the same instant). This also re-syncs this
            # process's own in-memory cache to the file's current real
            # state, since other machines' rows may have synced down too.
            verify = SharedLog()
            verify.load(self.path)
            newly_confirmed = [r for r in to_write if r["sync_id"] in verify.known_sync_ids]
            self.__dict__.update(verify.__dict__)

            return already_present + newly_confirmed
        except Exception:  # noqa: BLE001 -- a shared file we don't fully control must never crash the flush loop
            return []
        finally:
            _release_lock(self.path)

    def _pending_to_row(self, pending_row):
        """db.py's pending_sync row shape -> this module's own row-dict
        shape (FIELDS' key names). kind distinguishes a field-mapping
        confirmation from a value-map confirmation the same way the old
        append()/append_value_map() split used to -- a value_map row leaves
        desc blank and vice versa, matching FIELDS' own convention."""
        is_value_map = pending_row["kind"] == "value_map"
        return {
            "targetDatabase": pending_row["target_db"],
            "sourceSystem": pending_row["source_system"],
            "sourceField": pending_row["field_name"],
            "targetTable": pending_row["target_table"],
            "targetField": pending_row["target_field"],
            "confirmedAt": pending_row["confirmed_at"],
            "confirmedBy": pending_row["confirmed_by"],
            "desc": "" if is_value_map else (pending_row.get("desc") or ""),
            "valueMap": (pending_row.get("value_map") or "") if is_value_map else "",
            "syncId": pending_row["sync_id"],
        }

    def _extract_sync_ids(self, ws):
        """SyncIDs already present in a loaded (writable) worksheet -- reads
        directly off `ws` rather than a second SharedLog.load() call, since
        flush_pending already has this exact worksheet open."""
        if ws.max_row < 1:
            return set()
        header = [str(c.value).strip() if c.value else "" for c in ws[1]]
        if "SyncID" not in header:
            return set()
        col_idx = header.index("SyncID") + 1  # openpyxl columns are 1-indexed
        return {
            row[0] for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx, values_only=True)
            if row[0]
        }

    def _ensure_header_columns(self, ws):
        """Extends an existing header with any FIELDS columns it predates
        (e.g. Description, added after this log already had real rows) --
        always at the next free column, never touching existing header
        cells or shifting already-written data."""
        existing = [c.value for c in ws[1]]
        next_col = len(existing) + 1
        for header, _ in FIELDS:
            if header not in existing:
                ws.cell(row=1, column=next_col, value=header)
                existing.append(header)
                next_col += 1

    def _prepare_sheet(self, ws):
        if ws.max_row <= 1 and ws.cell(1, 1).value is None:
            # A brand-new (or truly empty, like the shared file starts out)
            # worksheet still reports max_row=1 for a phantom blank row, which
            # makes the *first* appended row land on row 2 instead of row 1 --
            # openpyxl quirk, verified directly. delete_rows resets that.
            ws.delete_rows(1, 1)
            ws.title = SHEET_NAME
            ws.append([header for header, _ in FIELDS])
        else:
            self._ensure_header_columns(ws)


def _lock_path(path):
    return path + ".synclock"


def _acquire_lock(path):
    """Best-effort advisory lock via a sidecar file -- NOT a real lock (see
    module docstring: OneDrive syncs each machine's own local replica of
    this folder asynchronously, so a lock file written here has no
    immediate effect on another machine's view of it until it syncs down).
    Narrows the collision window for the common case without pretending to
    guarantee exclusivity. Returns True if acquired (caller must call
    _release_lock when done), False if another writer still looks active
    after a few short waits -- the caller should back off and let the next
    scheduled flush retry, not block indefinitely."""
    lock_path = _lock_path(path)
    for _attempt in range(_LOCK_WAIT_ATTEMPTS):
        if os.path.exists(lock_path):
            age = time.time() - os.path.getmtime(lock_path)
            if age > _LOCK_STALE_SECONDS:
                try:
                    os.remove(lock_path)  # orphaned from a crashed writer, not an active one -- clear it and proceed
                except OSError:
                    pass
            else:
                time.sleep(_LOCK_WAIT_SECONDS)
                continue
        try:
            # O_CREAT | O_EXCL is an atomic create-if-not-exists at the OS
            # level -- stronger than a separate exists-check-then-write,
            # which two threads/processes could both pass at once.
            fd = os.open(lock_path, os.O_CREAT | os.O_EXCL | os.O_WRONLY)
            with os.fdopen(fd, "w", encoding="utf-8") as f:
                f.write(f"{socket.gethostname()} {getpass.getuser()} "
                        f"{datetime.datetime.now(datetime.timezone.utc).isoformat()}\n")
            return True
        except FileExistsError:
            time.sleep(_LOCK_WAIT_SECONDS)
    return False


def _release_lock(path):
    try:
        os.remove(_lock_path(path))
    except OSError:
        pass  # best-effort -- a failure to delete just leaves the lock looking "held" until it ages past _LOCK_STALE_SECONDS


SHARED = SharedLog()


def init():
    """Call once at process startup. Safe to call again to force a re-read."""
    SHARED.load(SHARED_XLSX_PATH)
    return SHARED
