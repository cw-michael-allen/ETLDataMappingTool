"""
Turns a real CaseWorthy Form XML export (<Ecm><Forms>/<Tables>/<Lists>...,
the "Export Form" feature in CaseWorthy admin) into the rows for a Field
Definition sheet -- ColumnName/TableName/FieldLabel/DataType/FormElementType/
Required/ListID/CharacterMaxLength/Comments -- matching the layout of the
custom-form sheets (WFDEmployment, WFDCheckIn) in the reference examples
Michael provided. See poc/form_xml.py for the generic structural preview
this builds on top of; that module still runs first and independently, so
an export this can't confidently interpret still gets *some* useful preview.

extract_form_templates() does the parsing; build_field_definition_workbook()
and build_staging_excel_workbook() (bottom of this file) turn its output into
the two downloadable .xlsx deliverables themselves, one sheet per form found.

Design decisions below are Michael's calls (2026-08-11), made reviewing a
real sample export (CaseWorthy_Export_Form_1000000013..., "Kiosk Intake
Review", Delaware People in Need) rather than guessed at:

- Column selection is FormElements-only: only fields actually placed on an
  exported form, not every column of the custom table backing it (a table
  can have far more columns than any one form exposes -- this sample's
  XKioskIntake table has ~50 columns, only 9 are on this form). <Tables> is
  used purely to enrich a matching column with its real SQL data type/length.
- A FormElement whose table isn't described in this export's own <Tables>
  section falls back to this repo's own base CaseWorthy schema
  (reference/target_schema_full.json, via schema_rules.load_schema) --
  added 2026-08-11 after a real "Add Client Demographics" export showed why
  the export-only lookup wasn't enough: CaseWorthy's Form export only
  includes <Tables> metadata for genuinely CUSTOM columns/tables (one
  export's <Tables> block described exactly one custom column added to
  Client, plus one fully-custom table -- nothing else), never redundantly
  re-describing its own standard schema (Client.FirstName, SSN, Gender,
  etc.), which this repo already has from the field-mapping side of the
  tool. Read-only reference lookup -- doesn't fold Create Template's own
  custom fields into that schema (Michael's call, 2026-08-11: keep Create
  Template decoupled from the mapping flow). A field that resolves in
  neither place (e.g. Family, FamilyMember -- tables outside both this
  export's <Tables> and the 28-table base schema) stays a genuine,
  reported gap, not a guess.
- System/audit columns (CreatedBy, CreatedDate, etc.) are NOT filtered out
  by name -- tried that first, but the sample export disproved it: its
  CreatedDate field is Usage="3" (a normal, visible field with its own
  "Created Date" label), deliberately placed on the review form for a human
  to see, not a throwaway system column. FormElements-only selection already
  means a human curated this list; a name-based filter on top of that would
  second-guess a deliberate choice. The reference examples agree -- WFDCheckIn's
  own Hidden Field rows (XCheckInID, ClientID) are kept as real rows too, just
  labeled Hidden Field, never dropped. So the only exclusion is DataInput="None"
  (a button/link/search box -- never real data at all).
- FormElementType comes from reference/cw_element_types.json, the real,
  complete ElementTypeID -> control-type lookup Michael provided directly
  (2026-08-11) -- not a guess. The one piece that IS inferred from the
  sample export rather than sourced: Usage="4" was observed marking a field
  "Hidden Field" even when its own ElementTypeID was "1" (Text Box) --
  CaseWorthy's Form Designer apparently lets a normal control be flagged
  hidden without switching its ElementTypeID. See HIDDEN_USAGE below.
- Decode values (the Comments column, for a ListID-backed field) come from
  this export's own embedded <Lists> block, not the baseline
  reference/cw_list_values.json registry -- an export's <Lists> section is
  self-contained and covers org-specific custom lists (the 1,000,000+ range)
  the baseline registry can't resolve at all.
- LinkedTo (added 2026-08-13, a new attribute row/column): whether a field's
  own (table, column) is a real foreign key elsewhere, e.g.
  ClientAddress.ClientID -> Client.EntityID -- this is what tells a
  consultant an ID in one sheet is the same thing as an ID in another,
  without which two sheets full of numbers can look unrelated even when
  they're not. Sourced from reference/cw_foreign_keys.json. A field with no
  matching entry has no known link, reported as such, never guessed at from
  column naming alone (an "*ID"-named column is not, by itself, evidence of
  anything) -- e.g. FamilyMember.OrgGroupID and
  EntityContactPreference.DeletedBy/LastModifiedBy/CreatedFormID have no
  entry here because the live database confirms they aren't real FK
  constraints, not because nobody checked.
- Synthesized join columns (both downloads, plus the browser preview since
  2026-08-12) come in two kinds, because joining two sheets needs a real
  column on BOTH sides: a child table's own FK column (_needed_link_columns,
  the original 2026-08-13 behavior) AND the parent/referenced table's own
  primary key column(s) (_needed_pk_columns, added 2026-08-12 after Michael
  pointed out a child's synthesized FK column is useless if the parent sheet
  never surfaces the column it's actually pointing at -- e.g. Client's own
  EntityID not being a field on any uploaded form). Sourced from
  reference/cw_primary_keys.json -- same trust rule as LinkedTo, never
  guessed from a column being named "*ID". Composite primary keys
  synthesize every missing column, not just one; a join can't work on half
  a key. Whether a join's *target* table counts as "present" is decided
  from the full export's own tables, never from which fields the customer
  currently has checked (_group_by_table_with_links' all_rows parameter) --
  found necessary 2026-08-14 after EntityContactPreference.EntityID ->
  Entity.EntityID vanished from a real download: Entity's own EntityID/
  EntityTypeID are both Hidden Fields the form never marks Required, so the
  app's default "Required only" starting selection left Entity with zero
  selected fields, and the old per-selection presence check treated that as
  "Entity isn't in this export." A parent table with no selected fields of
  its own now still gets its own sheet, containing just its synthesized
  primary key, rather than the join silently dropping after the preview
  banner already promised it.
- Third DataType fallback tier (added 2026-08-14): reference/cw_physical_columns.json,
  a full CaseWorthy database column export (every schema, table, and
  column), consulted only when neither an export's own <Tables> section nor
  the curated 28-table base CaseWorthy schema resolves a field. This is
  what closes the "table-alias gap" flagged after the base-schema fallback
  shipped: a Form export's FormElement DataInput references CaseWorthy's
  live *physical* table names (Entity, EntityContactPreference,
  ClientAddress, Family, FamilyMember, ...), which don't match the curated
  schema's own names (Client, AddressHistory) for the same conceptual data
  -- so a field on one of those physical tables stayed unresolved even with
  the base-schema fallback in place. This tier only ever supplies raw SQL
  type/length, never required/decode/ListID richness (the physical export
  has none of that -- it's a column dump, not an ETL-authored rule set),
  and only ever looks at the dbo schema, since a Form export's own physical
  table names are always dbo's even though the registry itself keeps every
  schema in the source export.

reference/cw_foreign_keys.json, reference/cw_primary_keys.json, and
reference/cw_physical_columns.json are all built by
tools/build_cw_baseline_schema.py from one consolidated, database-wide
schema+key export (provided directly by Michael, 2026-08-14) -- a single
source of these three facts, not three separate exports each with their own
staleness. See that script's own docstring for the source format and the
one-export-per-fact trust rule this repo holds all of them to.
"""

import json
import os
import re
import xml.etree.ElementTree as ET

import form_xml
import schema_rules

REFERENCE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "reference")
ELEMENT_TYPES_PATH = os.path.join(REFERENCE_DIR, "cw_element_types.json")
FOREIGN_KEYS_PATH = os.path.join(REFERENCE_DIR, "cw_foreign_keys.json")
PRIMARY_KEYS_PATH = os.path.join(REFERENCE_DIR, "cw_primary_keys.json")
PHYSICAL_COLUMNS_PATH = os.path.join(REFERENCE_DIR, "cw_physical_columns.json")
_ELEMENT_TYPES_CACHE = None
_BASE_SCHEMA_COLUMNS_CACHE = None
_FOREIGN_KEYS_CACHE = None
_PRIMARY_KEYS_CACHE = None
_PHYSICAL_COLUMNS_CACHE = None

_TEXT_MAX_RE = re.compile(r"Text \(max (\d+)\)")


def load_element_types():
    global _ELEMENT_TYPES_CACHE
    if _ELEMENT_TYPES_CACHE is None:
        if os.path.exists(ELEMENT_TYPES_PATH):
            with open(ELEMENT_TYPES_PATH, encoding="utf-8") as f:
                _ELEMENT_TYPES_CACHE = json.load(f)
        else:
            _ELEMENT_TYPES_CACHE = {}
    return _ELEMENT_TYPES_CACHE


def load_foreign_keys():
    global _FOREIGN_KEYS_CACHE
    if _FOREIGN_KEYS_CACHE is None:
        if os.path.exists(FOREIGN_KEYS_PATH):
            with open(FOREIGN_KEYS_PATH, encoding="utf-8") as f:
                _FOREIGN_KEYS_CACHE = json.load(f)
        else:
            _FOREIGN_KEYS_CACHE = {}
    return _FOREIGN_KEYS_CACHE


def load_primary_keys():
    global _PRIMARY_KEYS_CACHE
    if _PRIMARY_KEYS_CACHE is None:
        if os.path.exists(PRIMARY_KEYS_PATH):
            with open(PRIMARY_KEYS_PATH, encoding="utf-8") as f:
                _PRIMARY_KEYS_CACHE = json.load(f)
        else:
            _PRIMARY_KEYS_CACHE = {}
    return _PRIMARY_KEYS_CACHE


def load_physical_columns():
    global _PHYSICAL_COLUMNS_CACHE
    if _PHYSICAL_COLUMNS_CACHE is None:
        if os.path.exists(PHYSICAL_COLUMNS_PATH):
            with open(PHYSICAL_COLUMNS_PATH, encoding="utf-8") as f:
                _PHYSICAL_COLUMNS_CACHE = json.load(f)
        else:
            _PHYSICAL_COLUMNS_CACHE = {}
    return _PHYSICAL_COLUMNS_CACHE


def _linked_to(table_name, column_name):
    """'Client.EntityID'-style string when (table_name, column_name) is a
    real foreign key (reference/cw_foreign_keys.json), else None -- see this
    module's own docstring for why this is never guessed from column
    naming."""
    entry = load_foreign_keys().get(f"{table_name.lower()}.{column_name.lower()}")
    if not entry:
        return None
    return f"{entry['referencedTable']}.{entry['referencedColumn']}"


def _base_schema_columns():
    """(table_lower, field_lower) -> schema row, from schema_rules' own
    CaseWorthy schema -- see this module's docstring for why this fallback
    exists. schema_rules.load_schema() already re-reads+re-enriches the
    schema file fresh each call, so this cache is just this module's own
    keying of that same data, not a second source of truth."""
    global _BASE_SCHEMA_COLUMNS_CACHE
    if _BASE_SCHEMA_COLUMNS_CACHE is None:
        _BASE_SCHEMA_COLUMNS_CACHE = {
            (row["table"].lower(), row["field"].lower()): row
            for row in schema_rules.load_schema("CaseWorthy")
        }
    return _BASE_SCHEMA_COLUMNS_CACHE

# Same "Table.Column" convention file_import.py already parses for Advanced-
# mode headers -- duplicated rather than imported, matching this repo's own
# precedent (see transform_draft.py's _list_id_suffix) of not reaching across
# module boundaries for an underscore-prefixed helper.
_TABLE_COLUMN_RE = re.compile(r"^\s*([^.]+?)\s*\.\s*(.+?)\s*$")

# Usage="4" was observed on both of this sample's passthrough ID fields
# (XKioskIntakeID, X_ClientID) -- both otherwise ElementTypeID="1", same as
# ordinary visible text boxes -- so Usage, not ElementTypeID, is what marks a
# field "Hidden Field" and overrides whatever reference/cw_element_types.json
# would otherwise say for that ElementTypeID.
HIDDEN_USAGE = {"4"}

STRING_TYPES = {"nvarchar", "varchar", "char", "nchar"}


class CreateTemplateError(Exception):
    """Raised for any problem turning a Form export into template rows; message is user-facing."""


def _local(tag):
    return tag.rsplit("}", 1)[-1] if isinstance(tag, str) and "}" in tag else tag


def _parse_lists(root):
    """ListID (str) -> {"name": str|None, "values": [[code,label],...]},
    scoped to this export only -- see module docstring for why this is
    preferred over the baseline registry."""
    lists = {}
    for list_el in root.iter():
        if _local(list_el.tag) != "List":
            continue
        list_id = list_el.get("ListID")
        if not list_id:
            continue
        name_el = next((c for c in list_el if _local(c.tag) == "ListName"), None)
        items_el = next((c for c in list_el if _local(c.tag) == "ListItems"), None)
        values = []
        if items_el is not None:
            for item in items_el:
                if _local(item.tag) != "ListItem":
                    continue
                label_el = next((c for c in item if _local(c.tag) == "ListLabel"), None)
                code = item.get("ListValue")
                label = label_el.text if label_el is not None else None
                if code is not None and label is not None:
                    values.append([code, label])
        lists[list_id] = {"name": name_el.text if name_el is not None else None, "values": values}
    return lists


def _parse_table_columns(root):
    """(table_lower, column_lower) -> {dataType, length, isNullable,
    isPrimaryKey, isIdentity, foreignKeyTable}, from this export's own
    <Tables> section -- the real physical schema for whatever custom
    table(s) this export's form(s) are backed by."""
    columns = {}
    for table_el in root.iter():
        if _local(table_el.tag) != "Table":
            continue
        table_name = table_el.get("TableName")
        cols_el = next((c for c in table_el if _local(c.tag) == "Columns"), None)
        if not table_name or cols_el is None:
            continue
        for col in cols_el:
            if _local(col.tag) != "Column":
                continue
            col_name = col.get("ColumnName")
            if not col_name:
                continue
            fk_el = next((c for c in col if _local(c.tag) == "ForeignKey"), None)
            columns[(table_name.lower(), col_name.lower())] = {
                "dataType": col.get("DataType"),
                "length": col.get("Length"),
                "isNullable": col.get("IsNullable") == "1",
                "isPrimaryKey": col.get("IsPrimaryKey") == "1",
                "isIdentity": col.get("IsIdentity") == "1",
                "foreignKeyTable": fk_el.get("TableName") if fk_el is not None else None,
            }
    return columns


def _character_max_length(data_type, length):
    if not data_type or data_type.lower() not in STRING_TYPES or not length:
        return None
    try:
        n = int(length)
    except ValueError:
        return None
    if n < 0:
        return "MAX"  # SQL Server's nvarchar(MAX)/varchar(MAX) convention
    return n


def _form_element_type(usage, element_type_id):
    if usage in HIDDEN_USAGE:
        return "Hidden Field"
    entry = load_element_types().get(element_type_id)
    return entry["description"] if entry else f"Unknown (ElementTypeID={element_type_id})"


def _comments_for_list(list_id, lists, base_row=None):
    entry = lists.get(list_id) if list_id else None
    pairs = entry["values"] if entry and entry.get("values") else None
    if not pairs and base_row is not None:
        # This export's own <Lists> block has been comprehensive in every
        # sample seen so far (it covers every ListID any FormElement here
        # actually references), so this only matters if a future export
        # omits one -- falls back to the base schema's own decode/
        # decodeValues, same source _decode_mismatch/target_value_pairs use.
        pairs = schema_rules.target_value_pairs(base_row)
    if not pairs:
        return None
    # "{code} - {label}" newline-joined matches the reference Field
    # Definition examples' own Comments convention exactly (e.g. WFDEmployment's
    # "1 - Full Time\n5 - Part Time") -- deliberately not schema_rules.py's
    # "code=label, code=label" SQL-export decode format, a different
    # convention for a different document.
    return "\n".join(f"{code} - {label}" for code, label in pairs)


def _base_schema_lookup(table_name, column_name):
    """Fallback col_meta-equivalent from the base CaseWorthy schema, shaped
    just enough like table_columns' own entries for the caller to use either
    interchangeably. dataType is the schema's own `type` annotation (e.g.
    "Text (max 512)", "List", "Date") -- a different vocabulary than this
    export's own <Tables> SQL types ("nvarchar", "int"), but never translated
    into a SQL type that wasn't actually confirmed anywhere."""
    row = _base_schema_columns().get((table_name.lower(), column_name.lower()))
    if not row:
        return None
    type_str = row.get("type") or ""
    m = _TEXT_MAX_RE.match(type_str)
    return {
        "dataType": type_str,
        "characterMaxLength": int(m.group(1)) if m else None,
        "listId": row.get("listId"),
        "row": row,
    }


def _physical_schema_lookup(table_name, column_name, schema="dbo"):
    """Third and last DataType fallback -- see this module's own docstring
    for why this exists (closing the table-alias gap the base-schema
    fallback above couldn't). Only raw SQL type/length, shaped just enough
    like table_columns' own entries for the caller to use interchangeably;
    no required/decode/ListID richness, since reference/cw_physical_columns.json
    is a plain column export, not an ETL-authored rule set. Scoped to dbo by
    default -- a Form export's own physical table names are always dbo's,
    even though the registry itself keeps every schema from the source
    export (Michael's call, 2026-08-12: don't pre-filter the registry, scope
    at lookup time instead, same reasoning as build_cw_physical_columns.py's
    own docstring)."""
    row = load_physical_columns().get(f"{table_name.lower()}.{column_name.lower()}")
    if not row or row["schema"].lower() != schema.lower():
        return None
    return {
        "dataType": row["dataType"],
        "characterMaxLength": _character_max_length(row["dataType"], row["characterMaxLength"]),
    }


def extract_form_templates(raw_bytes):
    """raw_bytes -> list of {"formName", "formDisplayName", "rows": [...],
    "unresolvedColumns": [...]}, one entry per <Form> found.

    rows: [{"columnName", "tableName", "fieldLabel", "dataType",
            "formElementType", "required", "listId", "characterMaxLength",
            "linkedTo", "comments"}], in the same order fields appear in the
    export. linkedTo is a "Table.Column" string when this field is a real
    foreign key (reference/cw_foreign_keys.json), else None.

    unresolvedColumns: "Table.Column" strings for fields whose (table, column)
    isn't described in this export's own <Tables> section, the curated base
    CaseWorthy schema, or the dbo physical-column export (see
    _physical_schema_lookup) -- included so the field still shows up (a real
    field on the form) but with an honest gap on
    DataType rather than a guess.

    Raises CreateTemplateError if the file isn't a recognizable CaseWorthy
    Form export at all -- reuses form_xml's size/DOCTYPE/malformed-XML
    guards exactly, so both parsers agree on what's safe to even attempt.
    """
    if len(raw_bytes) > form_xml.MAX_UPLOAD_BYTES:
        limit_mb = form_xml.MAX_UPLOAD_BYTES // (1024 * 1024)
        raise CreateTemplateError(f"That file is larger than expected for a form export (limit {limit_mb} MB).")
    head = raw_bytes[:1024].lstrip()
    if b"<!DOCTYPE" in head.upper():
        raise CreateTemplateError("That file declares a DOCTYPE, which isn't accepted here.")
    try:
        root = ET.fromstring(raw_bytes)
    except ET.ParseError as e:
        raise CreateTemplateError(f"That doesn't look like valid XML: {e}")

    if _local(root.tag) != "Ecm":
        raise CreateTemplateError(
            "That doesn't look like a CaseWorthy Form export (expected an <Ecm> root element)."
        )

    lists = _parse_lists(root)
    table_columns = _parse_table_columns(root)

    forms_out = []
    for form_el in root.iter():
        if _local(form_el.tag) != "Form":
            continue
        form_elements_el = next((c for c in form_el if _local(c.tag) == "FormElements"), None)
        rows = []
        unresolved = []
        for fe in (form_elements_el or []):
            if _local(fe.tag) != "FormElement":
                continue
            data_input = fe.get("DataInput")
            if not data_input or data_input == "None":
                continue  # a button/link/search box, not a data field
            m = _TABLE_COLUMN_RE.match(data_input)
            if not m:
                continue
            table_name, column_name = m.group(1), m.group(2)

            label_el = next((c for c in fe if _local(c.tag) == "Label"), None)
            field_label = label_el.text if label_el is not None and label_el.text else column_name

            list_id = fe.get("ListID")
            base_row = None
            col_meta = table_columns.get((table_name.lower(), column_name.lower()))
            if col_meta is not None:
                data_type = col_meta["dataType"]
                char_max = _character_max_length(data_type, col_meta["length"])
            else:
                fallback = _base_schema_lookup(table_name, column_name)
                if fallback is not None:
                    data_type = fallback["dataType"]
                    char_max = fallback["characterMaxLength"]
                    base_row = fallback["row"]
                    if not list_id and fallback["listId"] is not None:
                        list_id = str(fallback["listId"])
                else:
                    physical = _physical_schema_lookup(table_name, column_name)
                    if physical is not None:
                        data_type = physical["dataType"]
                        char_max = physical["characterMaxLength"]
                    else:
                        data_type = None
                        char_max = None
                        unresolved.append(f"{table_name}.{column_name}")

            rows.append({
                "columnName": column_name,
                "tableName": table_name,
                "fieldLabel": field_label,
                "dataType": data_type,
                "formElementType": _form_element_type(fe.get("Usage"), fe.get("ElementTypeID")),
                "required": "Required" if fe.get("Required") == "1" else "Optional",
                "listId": list_id,
                "characterMaxLength": char_max,
                "linkedTo": _linked_to(table_name, column_name),
                "comments": _comments_for_list(list_id, lists, base_row),
            })

        forms_out.append({
            "formName": form_el.get("FormName"),
            "formDisplayName": form_el.get("FormDisplayName"),
            "rows": rows,
            "unresolvedColumns": unresolved,
        })

    if not forms_out:
        raise CreateTemplateError("No <Form> definitions found in that export.")
    return forms_out


# --- .xlsx generation -------------------------------------------------------
# openpyxl is an optional dependency (poc/requirements-optional.txt) -- the
# one deliberate exception to this repo's pure-stdlib rule, same reasoning as
# shared_mappings.py's own lazy import: a hand-rolled .xlsx *writer* is a much
# bigger, corruption-prone undertaking than file_import.py's read-only header
# parser. Mirrors shared_mappings._load_openpyxl's lazy-load pattern exactly
# rather than importing it, per this repo's own precedent (see
# transform_draft.py's _list_id_suffix) of not reaching across module
# boundaries for a private helper.
_openpyxl = None
_openpyxl_unavailable = None

OPENPYXL_UNAVAILABLE_MESSAGE = "openpyxl isn't installed (run: pip install openpyxl, or see poc/requirements-optional.txt)"

FIELD_DEF_ATTRS = [
    ("ColumnName", "columnName"),
    ("TableName", "tableName"),
    ("FieldLabel", "fieldLabel"),
    ("DataType", "dataType"),
    ("FormElementType", "formElementType"),
    ("Required", "required"),
    ("ListID", "listId"),
    ("CharacterMaxLength", "characterMaxLength"),
    ("LinkedTo", "linkedTo"),
    ("Comments", "comments"),
]

_INVALID_SHEET_CHARS = set(':\\/?*[]')


def _load_openpyxl():
    global _openpyxl, _openpyxl_unavailable
    if _openpyxl is not None or _openpyxl_unavailable is not None:
        return _openpyxl
    try:
        import openpyxl
        _openpyxl = openpyxl
    except ImportError:
        _openpyxl_unavailable = OPENPYXL_UNAVAILABLE_MESSAGE
    return _openpyxl


def _safe_sheet_name(name, used):
    """Excel sheet names: 31 chars max, none of : \\ / ? * [ ], and unique
    within the workbook -- a second form with the same display name gets a
    "(2)" suffix rather than silently overwriting the first one's sheet."""
    cleaned = "".join(c for c in (name or "Sheet") if c not in _INVALID_SHEET_CHARS).strip() or "Sheet"
    cleaned = cleaned[:31]
    base, n = cleaned, 2
    while cleaned in used:
        suffix = f" ({n})"
        cleaned = base[:31 - len(suffix)] + suffix
        n += 1
    used.add(cleaned)
    return cleaned


# Colors below are literal sRGB hex, not theme+tint references, so the
# generated file looks the same regardless of what theme it opens under --
# extracted by resolving the ICS reference workbook's own theme+tint values
# (Michael's "3. Staging Database Field Definition - ICS v8.xlsx", 2026-08-12)
# against its actual theme1.xml color scheme:
#   tab color:        theme 5 (accent2)  @ tint 0.7999816888943144 -> FBE5D6
#                      (Excel's built-in "Orange, Accent 2, Lighter 80%")
#   row-label fill:    theme 7 (accent4) @ tint 0.7999816888943144 -> FFF2CC
#                      (used consistently for the attribute-label column
#                      across every pink-tab sheet checked: WFDEmployment,
#                      WFDCheckIn)
#   ColumnName fill:   a pastel blue highlighting the row that identifies
#                      each field-column -- the reference varies which exact
#                      accent it starts from per sheet (WFDEmployment:
#                      theme 8/accent5 @ 0.5999938962981048 -> B4C7E7;
#                      WFDCheckIn: theme 4/accent1 @ same tint -> B7D6F2),
#                      so rather than copy that inconsistency this always
#                      uses B4C7E7 for a uniform look across every
#                      generated sheet.
_TAB_COLOR_PINK = "FBE5D6"
_ROW_LABEL_FILL = "FFF2CC"
_COLUMN_NAME_FILL = "B4C7E7"


def _style_field_definition_sheet(ws, num_fields):
    """Applies the ICS reference's formatting: bold gold row-label column,
    bold blue ColumnName header row, thin grid borders, wrapped text, frozen
    panes so both the attribute labels and the field-identifying row stay
    visible while scrolling a wide sheet. See build_field_definition_workbook
    for where the actual color values came from. Imports openpyxl.styles
    itself -- only ever called after build_field_definition_workbook has
    already confirmed openpyxl is installed."""
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    wrap = Alignment(wrap_text=True, vertical="top")
    row_label_fill = PatternFill("solid", fgColor=_ROW_LABEL_FILL)
    column_name_fill = PatternFill("solid", fgColor=_COLUMN_NAME_FILL)
    bold = Font(bold=True)

    last_col = num_fields + 1  # +1 for the row-label column itself
    for row_idx in range(1, len(FIELD_DEF_ATTRS) + 1):
        for col_idx in range(1, max(last_col, 1) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = wrap
            if col_idx == 1:
                cell.font = bold
                cell.fill = row_label_fill
            elif row_idx == 1:  # the ColumnName row identifies each field-column
                cell.font = bold
                cell.fill = column_name_fill

    ws.column_dimensions["A"].width = 22
    for col_idx in range(2, last_col + 1):
        col_letter = get_column_letter(col_idx)
        # Widest of that field's own values, capped so one long Comments
        # entry doesn't blow out the whole column -- Excel still wraps text
        # past this width, it just doesn't try to fit it on one line.
        widest = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, len(FIELD_DEF_ATTRS) + 1)),
            default=12,
        )
        ws.column_dimensions[col_letter].width = min(40, max(14, widest + 2))

    ws.freeze_panes = "B2"  # keeps both the attribute-label column and the ColumnName row in view


def _selected_rows(form_templates, selections):
    """Flattens every form's rows (respecting selections' per-form row-index
    filtering, same as before) into one list, dropping which form each row
    came from -- the workbook builders below regroup by tableName instead,
    not by form (Michael's call, 2026-08-13: a tab is a table, not a form/
    module, so a cross-table form's fields belong on separate tabs, and two
    forms that both touch the same table belong on the same one)."""
    selected = []
    for form_idx, form in enumerate(form_templates):
        allowed = None
        if selections is not None and form_idx < len(selections) and selections[form_idx] is not None:
            allowed = set(selections[form_idx])
        selected.extend(r for i, r in enumerate(form["rows"]) if allowed is None or i in allowed)
    return selected


def _group_by_table(rows):
    """tableName -> [rows], in first-seen order (of both tables and rows
    within a table) -- not alphabetical, so a table's columns stay in the
    order they appeared on the form(s) they came from.

    Deduplicates on (tableName, columnName): the same physical column can be
    surfaced on more than one form (e.g. two different forms both showing
    Client.Gender) -- it's one column, so it belongs on that table's tab
    once, not once per form that happened to display it. First occurrence
    wins if the two forms ever disagree about its metadata."""
    grouped = {}
    seen = set()
    for row in rows:
        key = (row["tableName"], row["columnName"])
        if key in seen:
            continue
        seen.add(key)
        grouped.setdefault(row["tableName"], []).append(row)
    return grouped


_FOREIGN_KEYS_BY_TABLE_CACHE = None
_PRIMARY_KEYS_BY_TABLE_CACHE = None


def _foreign_keys_by_table():
    """table_lower -> [fk entries whose OWN column lives on that table] --
    an index over load_foreign_keys() for the "which of this table's columns
    are foreign keys" direction, as opposed to _linked_to's "is this exact
    (table, column) a foreign key" direction."""
    global _FOREIGN_KEYS_BY_TABLE_CACHE
    if _FOREIGN_KEYS_BY_TABLE_CACHE is None:
        by_table = {}
        for entry in load_foreign_keys().values():
            by_table.setdefault(entry["table"].lower(), []).append(entry)
        _FOREIGN_KEYS_BY_TABLE_CACHE = by_table
    return _FOREIGN_KEYS_BY_TABLE_CACHE


def _primary_keys_by_table():
    """table_lower -> [pk column names], from reference/cw_primary_keys.json
    -- a table with a composite primary key yields more than one column
    here, since a join can't work on only half a composite key."""
    global _PRIMARY_KEYS_BY_TABLE_CACHE
    if _PRIMARY_KEYS_BY_TABLE_CACHE is None:
        by_table = {}
        for entry in load_primary_keys().values():
            by_table.setdefault(entry["table"].lower(), []).append(entry["column"])
        _PRIMARY_KEYS_BY_TABLE_CACHE = by_table
    return _PRIMARY_KEYS_BY_TABLE_CACHE


def _link_addition_row(table_name, fk_entry):
    """A synthetic row for a foreign-key column that's real (sourced from
    reference/cw_foreign_keys.json) but was never a FormElement on the
    export's own form(s) -- see module docstring's LinkedTo section for why
    this exists: CaseWorthy's live form auto-populates these from session
    context (the current client/entity), so a human never had to place them
    as a field, but a downloaded multi-sheet template still needs them to
    join a child table's rows back to its parent. Shaped exactly like a real
    row so it flows through the same rendering/workbook code unchanged;
    Comments and FormElementType make its synthetic origin unmistakable
    rather than passing it off as something a form actually asked for."""
    return {
        "columnName": fk_entry["column"],
        "tableName": table_name,
        "fieldLabel": fk_entry["column"],
        "dataType": None,
        "formElementType": "Linking Key (not on the original form)",
        "required": "Required",
        "listId": None,
        "characterMaxLength": None,
        "linkedTo": f"{fk_entry['referencedTable']}.{fk_entry['referencedColumn']}",
        "comments": (
            f"Added automatically -- needed to join {table_name} rows back to "
            f"{fk_entry['referencedTable']}, but wasn't a field on the original form."
        ),
        "kind": "foreign-key",
    }


def _pk_addition_row(table_name, pk_column):
    """A synthetic row for a table's own primary-key column that's real
    (sourced from reference/cw_primary_keys.json) but was never a
    FormElement on the export's own form(s) -- needed because a *child*
    table's synthesized FK column (_link_addition_row) is useless unless the
    *parent* table it points at actually surfaces the column being pointed
    to. Same shape as _link_addition_row's rows so it flows through the same
    rendering/workbook code unchanged; linkedTo is left empty since this
    column doesn't point anywhere itself -- it's the anchor other tables'
    FKs join back to."""
    return {
        "columnName": pk_column,
        "tableName": table_name,
        "fieldLabel": pk_column,
        "dataType": None,
        "formElementType": "Primary Key (not on the original form)",
        "required": "Required",
        "listId": None,
        "characterMaxLength": None,
        "linkedTo": None,
        "comments": (
            f"Added automatically -- {table_name}'s own primary key, needed so other "
            f"sheets' linking columns can actually join back to a real row here."
        ),
        "kind": "primary-key",
    }


def _needed_link_columns(grouped, all_tables):
    """tableName -> [synthetic rows] for every foreign key that's real
    (sourced), whose target table is also present among `all_tables` (see
    _group_by_table_with_links -- the FULL export's own tables, independent
    of which fields the customer currently has selected), and that isn't
    already covered by one of that table's own rows. Self-references (a
    table linking to itself, e.g. a HoH pointer within Client) are skipped
    -- not a cross-*sheet* concern, which is what this exists for. Always
    computed the same way regardless of the customer's checkbox selections
    -- a join key isn't optional the way a data field is (Michael's call,
    2026-08-13) -- which is why this checks `all_tables`, not `grouped`'s
    own keys: a target table whose own fields all happen to be deselected
    (or, as found 2026-08-14, never marked Required in the form -- e.g.
    Entity's EntityID/EntityTypeID are both Hidden Fields the form never
    flags Required, so the app's default "Required only" selection leaves
    Entity with nothing selected) is still a real join target, not an
    absent one."""
    present_tables = {t.lower() for t in all_tables}
    additions = {}
    for table_name, rows in grouped.items():
        existing_cols = {r["columnName"].lower() for r in rows}
        for fk in _foreign_keys_by_table().get(table_name.lower(), []):
            ref_lower = fk["referencedTable"].lower()
            if ref_lower == table_name.lower() or ref_lower not in present_tables:
                continue
            if fk["column"].lower() in existing_cols:
                continue
            additions.setdefault(table_name, [])
            if any(a["columnName"].lower() == fk["column"].lower() for a in additions[table_name]):
                continue  # a second FK constraint on the same column -- one synthetic row is enough
            additions[table_name].append(_link_addition_row(table_name, fk))
    return additions


def _needed_pk_columns(grouped, all_tables):
    """tableName -> [synthetic PK rows] for every table in `all_tables`
    (the FULL export's own tables -- see _needed_link_columns for why this
    isn't just `grouped`'s own keys) that's the REFERENCED side of some
    present table's real foreign key (self-references excluded, same rule
    as _needed_link_columns) and whose own primary-key column(s) aren't
    already present among that table's own selected rows -- see module
    docstring: a child's synthesized FK column is dead weight if the parent
    sheet never surfaces the column it's pointing at. A referenced table
    that currently has zero selected fields of its own (found 2026-08-14:
    Entity, under the app's default "Required only" selection -- see
    _needed_link_columns) isn't yet a key in `grouped` at all; this
    introduces it as a new table/sheet containing just its synthesized
    primary key, rather than only ever appending to a table already
    present, since otherwise a child's FK column would point at a sheet
    that silently doesn't exist in the download. Missing composite-key
    columns are all added, not just one."""
    lower_to_actual = {t.lower(): t for t in all_tables}
    referenced_lower = set()
    for table_name in grouped:
        for fk in _foreign_keys_by_table().get(table_name.lower(), []):
            ref_lower = fk["referencedTable"].lower()
            if ref_lower != table_name.lower() and ref_lower in lower_to_actual:
                referenced_lower.add(ref_lower)

    additions = {}
    for ref_lower in referenced_lower:
        actual_table = lower_to_actual[ref_lower]
        existing_cols = {r["columnName"].lower() for r in grouped.get(actual_table, [])}
        missing = [pk for pk in _primary_keys_by_table().get(ref_lower, []) if pk.lower() not in existing_cols]
        if missing:
            additions[actual_table] = [_pk_addition_row(actual_table, pk) for pk in missing]
    return additions


def _needed_join_columns(grouped, all_tables):
    """tableName -> [synthetic rows] combining both halves of what a real
    join needs: the child table's own FK column (_needed_link_columns) and
    the parent table's own primary key column(s) (_needed_pk_columns) --
    a join can't work with only one side of it present. Primary-key rows
    are listed first (the table's own identity), then FK rows, matching
    every reference example's convention of ID/link columns reading first;
    deduplicated by column name in the rare case a table needs the same
    column name from both halves. all_tables: see _needed_link_columns --
    the full export's own tables, not just `grouped`'s (currently
    selected) ones."""
    fk_additions = _needed_link_columns(grouped, all_tables)
    pk_additions = _needed_pk_columns(grouped, all_tables)
    combined = {}
    for table_name in set(fk_additions) | set(pk_additions):
        rows, seen = [], set()
        for row in pk_additions.get(table_name, []) + fk_additions.get(table_name, []):
            key = row["columnName"].lower()
            if key in seen:
                continue
            seen.add(key)
            rows.append(row)
        combined[table_name] = rows
    return combined


def _group_by_table_with_links(rows, all_rows=None):
    """_group_by_table, plus synthetic join columns (see
    _needed_join_columns) prepended to each affected table -- ID/link
    columns read first in every reference example seen, so a synthesized
    one is placed the same way.

    all_rows: the FULL, unfiltered row set for this export -- defaults to
    `rows` itself when omitted, so the browser preview's own call (which
    already passes every row, unfiltered, since a join key isn't optional
    per-selection) is unchanged. The actual downloads (build_field_definition_workbook,
    build_staging_excel_workbook) pass `rows` as whatever the customer
    currently has selected but `all_rows` as everything, so a join target
    whose own fields are all deselected -- or, as found 2026-08-14, never
    marked Required in the form at all, so never auto-selected by the
    app's default "Required only" starting selection -- still gets counted
    as present and, if needed, gets its own synthesized-PK-only sheet
    (_needed_pk_columns) rather than silently vanishing from the real
    download after the preview banner already promised it."""
    grouped = _group_by_table(rows)
    all_tables = _group_by_table(rows if all_rows is None else all_rows).keys()
    for table_name, additions in _needed_join_columns(grouped, all_tables).items():
        grouped[table_name] = additions + grouped.get(table_name, [])
    return grouped


def compute_link_additions(form_templates):
    """Public wrapper for the browser preview: every synthetic join column
    that will be added at download time, regardless of checkbox selections
    (see _needed_join_columns), as a flat list of
    {"table", "column", "linkedTo", "kind"} -- kind is "foreign-key"
    (linkedTo set) or "primary-key" (linkedTo is None -- nothing to link to,
    it IS the anchor) -- so a customer sees this before they download, not
    after."""
    grouped = _group_by_table(_selected_rows(form_templates, None))
    return [
        {"table": table_name, "column": row["columnName"], "linkedTo": row["linkedTo"], "kind": row["kind"]}
        for table_name, rows in _needed_join_columns(grouped, grouped.keys()).items()
        for row in rows
    ]


def compute_link_additions_by_form(form_templates):
    """Same computation as compute_link_additions, re-attributed per form so
    the browser preview -- which stays organized by FORM rather than by
    table (Michael's call, 2026-08-12: keep the existing per-form checkbox
    UI, don't restructure the preview to match the downloads' per-table
    grouping) -- can splice each table's synthetic rows into every
    form-block that touches that table. Two forms touching the same table
    both see that table's additions; not deduplicated across forms, since
    each form renders its own independent preview table. Ignores
    selections, same as compute_link_additions -- a join key isn't
    optional. Returns a list aligned 1:1 with form_templates."""
    grouped = _group_by_table(_selected_rows(form_templates, None))
    additions_by_table = _needed_join_columns(grouped, grouped.keys())
    result = []
    for form in form_templates:
        tables_touched, seen_tables = [], set()
        for row in form["rows"]:
            if row["tableName"] not in seen_tables:
                seen_tables.add(row["tableName"])
                tables_touched.append(row["tableName"])
        form_additions = []
        for table_name in tables_touched:
            form_additions.extend(additions_by_table.get(table_name, []))
        result.append(form_additions)
    return result


def build_field_definition_workbook(form_templates, selections=None):
    """One sheet per TARGET TABLE (not per form -- see _group_by_table),
    transposed (columns=fields, rows=attributes) -- matching the reference
    WFDEmployment/WFDCheckIn Field Definition sheets' layout AND formatting
    (pink tab, bold gold row labels, bold blue ColumnName row, grid borders
    -- see _style_field_definition_sheet). Returns an openpyxl Workbook;
    raises CreateTemplateError if openpyxl isn't installed.

    selections: a list, one entry per form (same order as form_templates),
    of row indices into that form's own "rows" to include. Both this and
    build_staging_excel_workbook honor the same selections shape (Michael's
    call, 2026-08-12) -- None includes every row, unfiltered.

    Passes the full unfiltered rows to _group_by_table_with_links alongside
    the selected ones (found necessary 2026-08-14) so a join target whose
    own fields are all deselected -- or never marked Required in the form,
    so never auto-selected by the app's default selection -- still counts
    as present instead of silently dropping its synthesized primary key
    (and, transitively, any other table's link to it) from the download."""
    openpyxl = _load_openpyxl()
    if openpyxl is None:
        raise CreateTemplateError(_openpyxl_unavailable)

    grouped = _group_by_table_with_links(
        _selected_rows(form_templates, selections), _selected_rows(form_templates, None)
    )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()
    for table_name, table_rows in grouped.items():
        sheet_name = _safe_sheet_name(table_name or "Unknown", used_names)
        ws = wb.create_sheet(sheet_name)
        ws.sheet_properties.tabColor = _TAB_COLOR_PINK
        for row_idx, (label, key) in enumerate(FIELD_DEF_ATTRS, start=1):
            ws.cell(row=row_idx, column=1, value=label)
            for col_idx, row in enumerate(table_rows, start=2):
                value = row.get(key)
                ws.cell(row=row_idx, column=col_idx, value=value if value is not None else "")
        _style_field_definition_sheet(ws, len(table_rows))
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    return wb


def build_staging_excel_workbook(form_templates, selections=None):
    """One sheet per TARGET TABLE (not per form -- see _group_by_table): a
    single flat header row of ColumnName values, matching every other
    Staging Excel sheet in this app's convention (header-row-only, see
    file_import.py). Returns an openpyxl Workbook; raises
    CreateTemplateError if openpyxl isn't installed.

    selections: same shape as build_field_definition_workbook's own -- None
    (or a form with no corresponding entry) includes every row, unfiltered.
    Passes the full unfiltered rows to _group_by_table_with_links alongside
    the selected ones -- see build_field_definition_workbook's own
    docstring for why."""
    openpyxl = _load_openpyxl()
    if openpyxl is None:
        raise CreateTemplateError(_openpyxl_unavailable)

    grouped = _group_by_table_with_links(
        _selected_rows(form_templates, selections), _selected_rows(form_templates, None)
    )
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    used_names = set()
    for table_name, table_rows in grouped.items():
        sheet_name = _safe_sheet_name(table_name or "Unknown", used_names)
        ws = wb.create_sheet(sheet_name)
        for col_idx, row in enumerate(table_rows, start=1):
            ws.cell(row=1, column=col_idx, value=row["columnName"])
    if not wb.sheetnames:
        wb.create_sheet("Sheet1")
    return wb
