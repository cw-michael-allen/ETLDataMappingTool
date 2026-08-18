"""
Extracts CaseWorthy's real Master Migration Template -- the actual staging
workbook a customer's migration gets checked against, not schema_rules.py's
own broader per-field `required` flag (extracted from CaseWorthy's Excel-
import validation script, which flags plenty of technically-required fields
across all 28 curated tables that a typical migration never touches) -- into
reference/cw_master_template.json.

Why this exists (Michael, 2026-08-17): the Migration Readiness rollup
(poc/readiness.py) was flagging required-missing fields across the full
28-table schema_rules schema, which double-counts PK/FK/technical
requiredness that isn't what the Master Template actually asks a migration
to supply. This script makes the Master Template itself -- not the broader
technical schema -- the source of truth for readiness.py's required-field
check on its own tables.

Source: reference/Sample_Staging_With_Data.xlsx. Each real sheet is one
target table; a header cell's yellow fill (RGB FFFFFF00 -- the only fill
color used across every header row in this workbook, confirmed by direct
inspection) marks that column as required, no-fill (00000000) marks it
optional. The `XNewCustomTable` sheet is deliberately excluded -- it's an
illustrative example of adding a table OUTSIDE the Master Template (no
column in it is highlighted at all), not a real required-fields definition;
readiness.py treats any table outside this file's own `tables` list as
exactly that case, falling back to known PK/FK constraints instead.

Run:
    python tools/build_master_template.py [--source PATH]

Needs openpyxl (poc/requirements-optional.txt) -- a one-off dev-tool
dependency, not an app runtime one, same as this repo's other tools/*.py
scripts and reprocessing scripts (see root CLAUDE.md).
"""

import argparse
import json
import os
import sys

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DEFAULT_SOURCE_XLSX = os.path.join(REPO_ROOT, "reference", "Sample_Staging_With_Data.xlsx")
OUT_PATH = os.path.join(REPO_ROOT, "reference", "cw_master_template.json")
SCHEMA_PATH = os.path.join(REPO_ROOT, "reference", "target_schema_full.json")

REQUIRED_FILL_RGB = "FFFFFF00"

# Illustrative "here's how you'd add a custom table" example, not a real
# required-fields definition -- see module docstring.
EXCLUDED_SHEETS = {"XNewCustomTable"}


def _header_cells(ws):
    return next(ws.iter_rows(min_row=1, max_row=1))


def extract(wb):
    tables = []
    required_fields = {}
    all_fields = {}
    for sheet_name in wb.sheetnames:
        if sheet_name in EXCLUDED_SHEETS:
            continue
        ws = wb[sheet_name]
        required, all_cols = [], []
        for cell in _header_cells(ws):
            if cell.value is None:
                continue
            field = str(cell.value).strip()  # source has at least one trailing-space header ("HoHClientID ")
            if not field:
                continue
            all_cols.append(field)
            fill = cell.fill.fgColor.rgb if cell.fill and cell.fill.fgColor else None
            if fill == REQUIRED_FILL_RGB:
                required.append(field)
        if not all_cols:
            continue
        tables.append(sheet_name)
        all_fields[sheet_name] = all_cols
        required_fields[sheet_name] = required
    return {"tables": sorted(tables), "requiredFields": required_fields, "allFields": all_fields}


def _report_schema_disagreements(result):
    """Informational only, never blocking: where this Master Template's own
    required flag disagrees with schema_rules' broader validation-script-
    derived one for the same (table, field) -- worth a human glance, same
    "report disagreements for sign-off rather than silently picking a
    winner" convention as tools/extract_servtracker_schema.py, but nothing
    here overrides the other; they're deliberately two different concepts
    now (see module docstring)."""
    if not os.path.exists(SCHEMA_PATH):
        return
    with open(SCHEMA_PATH, encoding="utf-8") as f:
        schema_rows = json.load(f)
    schema_required = {
        (row["table"], row["field"]) for row in schema_rows if row.get("required")
    }
    diffs = []
    for table in result["tables"]:
        required_here = set(result["requiredFields"].get(table, []))
        for field in result["allFields"].get(table, []):
            here = field in required_here
            there = (table, field) in schema_required
            if here != there:
                diffs.append((table, field, here, there))
    if not diffs:
        print("  No disagreements with target_schema_full.json's own `required` flag.")
        return
    print(f"  {len(diffs)} field(s) where Master Template required-ness differs from "
          f"target_schema_full.json's own `required` flag (informational only -- neither wins):")
    for table, field, here, there in diffs:
        print(f"    {table}.{field}: Master Template required={here}, target_schema_full.json required={there}")


def main():
    ap = argparse.ArgumentParser(description=__doc__.strip().split("\n")[0])
    ap.add_argument("--source", default=DEFAULT_SOURCE_XLSX)
    args = ap.parse_args()

    if not os.path.exists(args.source):
        sys.exit(f"Source not found: {args.source}")

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl is required to run this script: pip install openpyxl")

    wb = openpyxl.load_workbook(args.source, data_only=True)
    result = extract(wb)

    with open(OUT_PATH, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, sort_keys=True)
        f.write("\n")

    total_required = sum(len(v) for v in result["requiredFields"].values())
    print(f"Wrote {OUT_PATH}: {len(result['tables'])} tables "
          f"({', '.join(result['tables'])}), {total_required} required field(s) total.")
    _report_schema_disagreements(result)


if __name__ == "__main__":
    main()
